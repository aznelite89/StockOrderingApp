"""Build the CSV + XLSX exports for the invoice-to-Excel tool.

Kept separate from invoice_app.py so the functions can be imported and
tested without Streamlit's module-level side effects (set_page_config etc.).
"""

from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from constants.invoice import (
    FIRST_DATA_ROW_EXCEL,
    HeaderField,
    LINE_COLUMN_ORDER,
    LineColumn,
    OUTPUT_SHEET_NAME,
    PRICE_AFTER_DISCOUNT_FORMULA,
    RRP_FORMULA_TEMPLATE,
    RULES_SHEET_NAME,
    SUM_TOTAL_BEFORE_DISCOUNT_FORMULA,
    TABLE_HEADER_ROW_EXCEL,
    TOTAL_BEFORE_DISCOUNT_FORMULA,
)
from utils.pdf_invoice import ParsedInvoice


def excel_serial(dt: datetime) -> int:
    """Excel date serial — days since 1899-12-30."""
    return (dt - datetime(1899, 12, 30)).days


def build_csv(parsed: ParsedInvoice, df: pd.DataFrame, customer_code: str) -> bytes:
    """Values-only CSV matching the Speirs sample layout."""
    buf = io.StringIO()
    inv_date = parsed.header[HeaderField.INVOICE_DATE]
    inv_date_serial = excel_serial(inv_date) if inv_date else ""

    writer_rows: list[list] = [
        [HeaderField.CUSTOMER_NAME, parsed.header[HeaderField.CUSTOMER_NAME]],
        [HeaderField.CUSTOMER_CODE, customer_code],
        [HeaderField.INVOICE_DATE, inv_date_serial],
        [HeaderField.ORDER_REFERENCE, parsed.header[HeaderField.ORDER_REFERENCE]],
        [],
        LINE_COLUMN_ORDER,
    ]
    for _, row in df.iterrows():
        writer_rows.append([row[col] for col in LINE_COLUMN_ORDER])

    pd.DataFrame(writer_rows).to_csv(buf, header=False, index=False)
    return buf.getvalue().encode("utf-8")


def build_xlsx(parsed: ParsedInvoice, df: pd.DataFrame,
               customer_code: str, rules_bytes: Optional[bytes] = None,
               on_warning=None) -> bytes:
    """Formula-bearing XLSX matching the Speirs sample layout.

    `rules_bytes` is the uploaded Pricing Rules workbook contents; the
    `Pricing Rules` tab from that workbook is copied into the output.
    `on_warning(msg)` is called for non-fatal issues (e.g. missing rules tab).
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        workbook = writer.book
        ws = workbook.add_worksheet(OUTPUT_SHEET_NAME)
        writer.sheets[OUTPUT_SHEET_NAME] = ws

        bold = workbook.add_format({"bold": True})
        date_fmt = workbook.add_format({"num_format": "dd/mm/yyyy"})
        money_fmt = workbook.add_format({"num_format": "#,##0.00"})
        percent_fmt = workbook.add_format({"num_format": "0%"})
        header_fmt = workbook.add_format(
            {"bold": True, "bg_color": "#D3D3D3", "border": 1}
        )

        inv_date = parsed.header[HeaderField.INVOICE_DATE]
        ws.write(0, 0, HeaderField.CUSTOMER_NAME, bold)
        ws.write(0, 1, parsed.header[HeaderField.CUSTOMER_NAME])
        ws.write(1, 0, HeaderField.CUSTOMER_CODE, bold)
        ws.write(1, 1, customer_code)
        ws.write(2, 0, HeaderField.INVOICE_DATE, bold)
        if inv_date is not None:
            ws.write_datetime(2, 1, inv_date, date_fmt)
        ws.write(3, 0, HeaderField.ORDER_REFERENCE, bold)
        ws.write(3, 1, parsed.header[HeaderField.ORDER_REFERENCE])

        # Table header on row 6 (xlsxwriter 0-indexed row 5).
        for col_idx, name in enumerate(LINE_COLUMN_ORDER):
            ws.write(TABLE_HEADER_ROW_EXCEL - 1, col_idx, name, header_fmt)

        first_data_row = FIRST_DATA_ROW_EXCEL - 1
        for i, (_, row) in enumerate(df.iterrows()):
            r0 = first_data_row + i
            r1 = r0 + 1  # 1-indexed Excel row number, used in formulas
            ws.write_number(r0, 0, int(row[LineColumn.LINE_NO]))
            ws.write_string(r0, 1, row[LineColumn.CODE])
            ws.write_string(r0, 2, row[LineColumn.DESCRIPTION] or "")
            ws.write_number(r0, 3, float(row[LineColumn.QTY]))
            ws.write_number(r0, 4, float(row[LineColumn.PRICE]), money_fmt)
            ws.write_formula(
                r0, 5,
                TOTAL_BEFORE_DISCOUNT_FORMULA.format(r=r1),
                money_fmt,
                float(row[LineColumn.TOTAL_BEFORE_DISCOUNT]),
            )
            ws.write_number(r0, 6, float(row[LineColumn.DISC]), percent_fmt)
            ws.write_formula(
                r0, 7,
                PRICE_AFTER_DISCOUNT_FORMULA.format(r=r1),
                money_fmt,
                float(row[LineColumn.PRICE_AFTER_DISCOUNT]),
            )
            ws.write_number(
                r0, 8, float(row[LineColumn.TOTAL_AFTER_DISCOUNT]), money_fmt
            )
            ws.write_formula(
                r0, 9, RRP_FORMULA_TEMPLATE.format(r=r1), None,
                row[LineColumn.RRP],
            )

        last_data_row = first_data_row + len(df)
        ws.write_formula(
            last_data_row, 5,
            SUM_TOTAL_BEFORE_DISCOUNT_FORMULA.format(
                first=first_data_row + 1, last=last_data_row
            ),
            money_fmt,
        )

        widths = [16, 15, 50, 8, 10, 18, 8, 18, 18, 14]
        for i, w in enumerate(widths):
            ws.set_column(i, i, w)
        ws.freeze_panes(TABLE_HEADER_ROW_EXCEL, 0)

        if rules_bytes:
            _copy_rules_tab(rules_bytes, writer, on_warning)

    return buf.getvalue()


def _copy_rules_tab(rules_bytes: bytes, writer: pd.ExcelWriter, on_warning) -> None:
    def warn(msg):
        if on_warning:
            on_warning(msg)

    try:
        rules_wb = load_workbook(io.BytesIO(rules_bytes), data_only=True)
    except Exception as exc:
        warn(f"Could not read the Pricing Rules workbook: {exc}")
        return
    if RULES_SHEET_NAME not in rules_wb.sheetnames:
        warn(
            f"Uploaded workbook does not contain a `{RULES_SHEET_NAME}` tab; "
            "the output will not include a copy of the rules."
        )
        return
    src = rules_wb[RULES_SHEET_NAME]
    workbook = writer.book
    target = workbook.add_worksheet(RULES_SHEET_NAME)
    writer.sheets[RULES_SHEET_NAME] = target
    bold = workbook.add_format({"bold": True})
    for row_idx, row in enumerate(src.iter_rows(values_only=True)):
        for col_idx, value in enumerate(row):
            if value is None:
                continue
            fmt = (
                bold if row_idx == 0 or (
                    isinstance(value, str) and value in {"Factor", "Rounding"}
                ) else None
            )
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                target.write_number(row_idx, col_idx, value, fmt)
            else:
                target.write(row_idx, col_idx, value, fmt)
    target.set_column(0, 0, 18)
    target.set_column(1, 1, 28)
